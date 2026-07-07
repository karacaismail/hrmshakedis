# -*- coding: utf-8 -*-
"""hrmshakedis v4.1 — sıfırdan: YAZILIM GELİŞTİRME teklifi (waterfall) + parametreler. 6 sayfa, Pareto.
WEB_PARAM protokolü: tüm değerler METİN (=""&x / TEXT(t,"yyyy-mm-dd")) — gviz karışık-tip null sorununu çözer."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.page import PageMargins
import datetime as _dt

NAVY="0F172A"; BLUE="1A56DB"; BLUE_L="EBF1FE"; ZEBRA="F6F8FB"
BORD="D8E0EA"; GREEN="047857"; RED="C81E1E"; AMBER="B45309"; MUTED="64748B"
OK_F="DEF7EC"; GRAY_H="334155"
TL='#,##0.00" TL"'; PCT='0.00%'; DT='DD.MM.YYYY'
F=lambda **k: Font(**{"name":"Roboto","size":12,**k})
thin=Side(style="thin", color=BORD)
BX=Border(left=thin,right=thin,top=thin,bottom=thin)
WRAP=Alignment(vertical="top", wrap_text=True)
CTR=Alignment(horizontal="center", vertical="center", wrap_text=True)

def sheet(wb,name,tab):
    ws=wb.create_sheet(name); ws.sheet_view.showGridLines=False; ws.sheet_properties.tabColor=tab
    return ws
def title(ws,txt,sub=None):
    ws["A1"]=txt; ws["A1"].font=F(size=17,bold=True,color=NAVY)
    if sub: ws["A2"]=sub; ws["A2"].font=F(color=MUTED)
    ws.row_dimensions[1].height=26
def hdr_row(ws,r,cols,start=1):
    for i,c in enumerate(cols):
        cell=ws.cell(row=r,column=start+i,value=c)
        cell.font=F(bold=True,color="FFFFFF"); cell.fill=PatternFill("solid",start_color=NAVY)
        cell.alignment=CTR; cell.border=BX
    ws.row_dimensions[r].height=30
def sect(ws,r,txt):
    c=ws.cell(row=r,column=1,value=txt); c.font=F(bold=True,size=13,color=BLUE); ws.row_dimensions[r].height=22
def body(ws,r0,rows,heights=None):
    for j,row in enumerate(rows):
        r=r0+j
        for i,v in enumerate(row):
            cell=ws.cell(row=r,column=1+i,value=v)
            cell.font=F(); cell.border=BX; cell.alignment=WRAP
            if j%2==1: cell.fill=PatternFill("solid",start_color=ZEBRA)
        if heights: ws.row_dimensions[r].height=heights
def widths(ws,m):
    for col,w in m.items(): ws.column_dimensions[col].width=w
def inp(ws,addr,val,fmt=None):
    c=ws[addr]; c.value=val; c.font=F(color=BLUE,bold=True)
    c.fill=PatternFill("solid",start_color=BLUE_L); c.border=BX
    if fmt: c.number_format=fmt
def out(ws,addr,formula,fmt=None):
    c=ws[addr]; c.value=formula; c.font=F(); c.border=BX
    if fmt: c.number_format=fmt
def lbl(ws,addr,txt,bold=False,color=NAVY):
    c=ws[addr]; c.value=txt; c.font=F(bold=bold,color=color); c.alignment=WRAP; c.border=BX

wb=Workbook(); wb.remove(wb.active)
G="'TEKLİF GİRDİ'!"; P="PARAMETRELER!"; W="'WATERFALL PLAN'!"

# ============ 1 OKU BENİ ============
ws=sheet(wb,"OKU BENİ",NAVY)
title(ws,"hrmshakedis — Yazılım Geliştirme Teklifi Çalışma Kitabı",
      "Sürüm 4.1 · 07.07.2026 · Müşteri: İnşaat firması · Waterfall proje tanımı · 6 sayfa (Pareto: müşterinin görecekleri + parametreler)")
rows=[
 ["Bu dosya nedir?","Asgari işçilik & hakediş uyum platformunun GELİŞTİRİLMESİ için yazılım teklifi kitabı: kapsam seçimi, waterfall faz planı, A4 yazdırılabilir YAZILIM GELİŞTİRME TEKLİF FORMU ve tüm değişken parametreler. Ürünün çözdüğü ihtiyaç, ekteki asgari işçilik tablosu ve SGK Genelge 2026/16'dır."],
 ["Sayfalar","TEKLİF GİRDİ: kimlik + kapsam seçimi (EVET/HAYIR) · TEKLİF FORMU: A4 çıktı · WATERFALL PLAN: 6 fazın tarihleri (otomatik) · PARAMETRELER: hız modeli, efor, fiyat, mevzuat mini · WEB_PARAM: web beslemesi (dokunmayın)."],
 ["Canlı bağ","GitHub Pages raporu ve teklif sayfası bu dosyanın Google Sheets kopyasındaki WEB_PARAM sayfasını okur; Sheets'te değişen her şey web'e ~60 sn içinde yansır. Adres: karacaismail.github.io/hrmshakedis"],
 ["Google Sheets","YALNIZCA şu adreste yayınlanır: docs.google.com/spreadsheets/d/1VXwPEh-eEIYXqRpxRO3MoN3cib-R5jBjkNRU-GPJIYw — Rutin değişiklikler artık CANLI yapılır: Claude + Zapier 'hrmshakedis sheet güncelle' becerisi hücreleri yerinde günceller. 'Dosya > İçe aktar > E-tabloyu değiştir' yalnız YAPISAL değişiklikte kullanılır. Kurallar: docs/data-map.md ve docs/guides/. Makro yok; formüller Sheets uyumlu."],
 ["Hız modeli","Tamamen vibecoding: AI kodu ultra hızlı yazar, insan doğrular. Takvim çarpanı = Amdahl × ECA ÷ geliştirici sayısı (LİNEER: 3 kişi → süre ÷ 3). PARAMETRELER §1'den yönetilir; fazlar, GA ve web otomatik ölçeklenir. Bedel, insan-eşdeğeri efora (adam-gün) dayanır; AI hızlanması takvimi kısaltır, bedeli değiştirmez."],
 ["Renk kodu","Mavi zeminli mavi yazı = SİZİN GİRECEĞİNİZ değer. Siyah = formül (dokunmayın)."],
]
hdr_row(ws,4,["Başlık","Açıklama"]); body(ws,5,rows,heights=56)
widths(ws,{"A":22,"B":140}); ws.freeze_panes="A5"

# ============ 2 TEKLİF GİRDİ ============
ws=sheet(wb,"TEKLİF GİRDİ",AMBER)
title(ws,"Yazılım Geliştirme Teklifi — Girdiler","Kimlik bilgilerini yazın, kapsamı EVET/HAYIR ile seçin. Form ve web anında güncellenir.")
sect(ws,4,"1) Kimlik")
kim=[("Müşteri","İnşaat firması"),("Proje adı","Asgari İşçilik & Hakediş Uyum Platformu"),
     ("Teklif no","SW-2026-001"),("Teklif tarihi",_dt.date(2026,7,7)),("Geçerlilik (gün)",30),
     ("Hazırlayan / Yüklenici","İsmail Karaca — bağımsız yazılım geliştirici · karacai@yandex.com")]
for j,(t,v) in enumerate(kim):
    r=5+j; lbl(ws,f"B{r}",t); inp(ws,f"C{r}",v)
ws["C8"].number_format=DT; ws["C9"].number_format="0"
sect(ws,12,"2) Kapsam seçimi (modüller)")
hdr_row(ws,13,["","Modül","Dahil?","Taban efor (adam-gün)","İçerik"],start=1)
mods=[
 ("Çekirdek altyapı & CI (zorunlu)","EVET","Excel'in hiç yapamadığını ekler: yedeklilik, kalite kapıları, iki saatte kurulabilirlik"),
 ("M1 · Ruhsat & alan bazlı hesap","EVET","RUHSAT HESAPLAMASI sayfasının ürün hâli; 2026/16 ortalama kuralı ve giriş denetimi"),
 ("M2 · İhale & sözleşme hesabı + kısmi araştırma","EVET","GENEL HESAPLAMA sayfasının ürün hâli; kırılan Etap-2 bağları ve mahsup akışı"),
 ("M3 · Aylık bildirim matrisi, İcmal & kota","EVET","ASGARİ İŞÇİLİK 1&2 matrisi + İCMAL panelinin ürün hâli; alt-taşeron kuralları, elle 'kalan adam/ay' biter"),
 ("M4 · Mevzuat parametre yönetimi","EVET","Formüllere gömülü 6,75 / 0,345 / 0,375 / 38,5 sabitleri tek tabloya taşınır; dönem kilidi"),
 ("M5 · Google Sheets köprüsü & veri teyidi","EVET","'Manuel elle yazılacak alan' ve 'TEYİT EDİLDİ' notları, denetimli giriş ve izlenen teyide dönüşür"),
 ("M6 · Taşeron sicili, raporlar & KVKK","EVET","TAŞERON sicil sayfasının ürün hâli; maskeleme, evrak takibi, dört denetim raporu"),
]
for j,(ad,sec,ic) in enumerate(mods):
    r=14+j
    lbl(ws,f"A{r}",f"K{j}") ; lbl(ws,f"B{r}",ad)
    inp(ws,f"C{r}",sec)
    out(ws,f"D{r}",f"=VLOOKUP(B{r},{P}$A$21:$B$27,2,FALSE)",'#,##0" a-g"')
    lbl(ws,f"E{r}",ic,color=MUTED)
    if j%2==1:
        for col in "ABDE": ws[f"{col}{r}"].fill=PatternFill("solid",start_color=ZEBRA)
dv=DataValidation(type="list",formula1='"EVET,HAYIR"',allow_blank=False); ws.add_data_validation(dv); dv.add("C15:C20")
lbl(ws,"C14b" if False else "F14","",False)
lbl(ws,"A22","Seçili toplam efor:",bold=True)
out(ws,"D22",f'=SUMPRODUCT((C14:C20="EVET")*D14:D20)','#,##0" adam-gün"'); ws["D22"].font=F(bold=True,color=BLUE)
lbl(ws,"E22","Çekirdek altyapı zorunludur (K0 = EVET kalmalı). Süre ve bedel bu toplamdan türetilir.",color=MUTED)
widths(ws,{"A":5,"B":40,"C":12,"D":20,"E":70})
ws.freeze_panes="A5"

# ============ 3 WATERFALL PLAN ============
ws=sheet(wb,"WATERFALL PLAN",GREEN)
title(ws,"Waterfall Faz Planı — Otomatik Tarihler","Süreler: seçili efor × zaman çarpanı × faz payı. PARAMETRELER değişince bu plan, form ve web otomatik güncellenir.")
hdr_row(ws,4,["Faz","Teslimatlar (Definition of Done)","Pay","Süre (gün)","Başlangıç","Bitiş"])
fazlar=[
 ("F1 · Analiz & Gereksinim","Onaylı gereksinim dokümanı; 12 altın senaryonun (genelge + müşteri verisi) kabul testi olarak dökümü; kapsam dondurma.",0.12),
 ("F2 · Tasarım","Mimari + veri modeli (12 veri nesnesi); API sözleşmesi; test planı; KVKK veri sınıflaması.",0.15),
 ("F3 · Geliştirme (vibecoding)","AI ajanlarıyla modül geliştirme; her PR'da CI kapıları; haftalık çalışan ara sürüm.",0.38),
 ("F4 · Test & Doğrulama","Altın dosya paritesi 12/12; entegrasyon + API sözleşme + güvenlik testleri; yük testi p95 < 300 ms.",0.18),
 ("F5 · UAT / Geçiş","Müşteri kabul testi (gerçek Defne verisiyle mutabakat, fark 0); eğitim; veri göçü provası.",0.10),
 ("F6 · Devreye Alma & Hypercare","Üretim yayını (GA); 2 hafta yakın izleme; runbook + devir dokümanları.",0.07),
]
for j,(ad,ts,pay) in enumerate(fazlar):
    r=5+j
    lbl(ws,f"A{r}",ad); lbl(ws,f"B{r}",ts,color=MUTED)
    out(ws,f"C{r}",pay,PCT)
    if j<5: out(ws,f"D{r}",f"=MAX(1,ROUND({P}$B$31*C{r},0))",'0" gün"')
    else:   out(ws,f"D{r}",f"={P}$B$31-SUM(D5:D9)",'0" gün"')
    out(ws,f"E{r}",f"={P}$B$33" if j==0 else f"=F{r-1}+1",DT)
    out(ws,f"F{r}",f"=E{r}+D{r}-1",DT)
    if j%2==1:
        for col in "ABCDEF": ws[f"{col}{r}"].fill=PatternFill("solid",start_color=ZEBRA)
    ws.row_dimensions[r].height=40
lbl(ws,"A12","TOPLAM / GA",bold=True)
out(ws,"D12",'=SUM(D5:D10)','0" gün"'); ws["D12"].font=F(bold=True)
out(ws,"F12","=F10",DT); ws["F12"].font=F(bold=True,color=BLUE)
lbl(ws,"A13","Tampon sonu (+%15)",bold=True); out(ws,"F13",f"={P}$B$35",DT)
lbl(ws,"A15","Waterfall notu",bold=True)
lbl(ws,"A16","Fazlar sıralı ilerler; bir faz kapanmadan diğeri başlamaz (kapı: teslimatların onayı). Vibecoding hızlanması takvimi kısaltır, kapıları kaldırmaz. Mevzuat değişikliği kapsam değişikliği sayılır ve F1'e döner.",color=MUTED)
ws.merge_cells("A16:F17"); ws["A16"].alignment=WRAP
widths(ws,{"A":30,"B":66,"C":9,"D":12,"E":14,"F":14})
ws.freeze_panes="A5"

# ============ 4 TEKLİF FORMU (A4) ============
ws=sheet(wb,"TEKLİF FORMU",AMBER)
ws.merge_cells("B2:E2"); ws["B2"]="YAZILIM GELİŞTİRME TEKLİF FORMU"
ws["B2"].font=F(size=16,bold=True,color=NAVY); ws["B2"].alignment=CTR; ws.row_dimensions[2].height=30
ws.merge_cells("B3:E3"); ws["B3"]='=""&'+G+'C6&" — özel yazılım geliştirme teklifi"'
ws["B3"].font=F(color=MUTED); ws["B3"].alignment=CTR
def band(r,txt):
    ws.merge_cells(f"B{r}:E{r}"); c=ws[f"B{r}"]; c.value=txt
    c.font=F(bold=True,color="FFFFFF"); c.fill=PatternFill("solid",start_color=NAVY); c.alignment=CTR
    ws.row_dimensions[r].height=24
band(5,"TARAFLAR VE KİMLİK")
lbl(ws,"B6","Teklif no"); out(ws,"C6",f"={G}C7")
lbl(ws,"D6","Tarih"); out(ws,"E6",f"={G}C8",DT)
lbl(ws,"B7","Müşteri"); out(ws,"C7",f"={G}C5")
lbl(ws,"D7","Geçerlilik"); out(ws,"E7",f'={G}C9&" gün"')
lbl(ws,"B8","Yüklenici"); out(ws,"C8",f"={G}C10"); ws.merge_cells("C8:E8")
lbl(ws,"B9","Proje"); out(ws,"C9",f"={G}C6"); ws.merge_cells("C9:E9")
band(11,"KAPSAM (SEÇİLİ MODÜLLER)")
for j in range(7):
    r=12+j; gr=14+j
    out(ws,f"B{r}",f'=IF({G}C{gr}="EVET","• "&{G}B{gr}&"  ("&TEXT({G}D{gr},"0")&" adam-gün)","")')
    ws.merge_cells(f"B{r}:E{r}"); ws[f"B{r}"].font=F()
band(20,"WATERFALL PLAN")
lbl(ws,"B21","Faz",bold=True); lbl(ws,"C21","Süre",bold=True); lbl(ws,"D21","Başlangıç",bold=True); lbl(ws,"E21","Bitiş",bold=True)
for j in range(6):
    r=22+j; wr=5+j
    out(ws,f"B{r}",f"={W}A{wr}")
    out(ws,f"C{r}",f'={W}D{wr}&""&" gün"' if False else f'=TEXT({W}D{wr},"0")&" gün"')
    out(ws,f"D{r}",f"={W}E{wr}",DT); out(ws,f"E{r}",f"={W}F{wr}",DT)
band(29,"TİCARİ KOŞULLAR")
lbl(ws,"B30","Toplam efor"); out(ws,"C30",f'=TEXT({G}D22,"#,##0")&" adam-gün"')
lbl(ws,"D30","Toplam süre"); out(ws,"E30",f'=TEXT({W}D12,"0")&" gün"')
lbl(ws,"B31","Teslim (GA)"); out(ws,"C31",f"={W}F12",DT)
lbl(ws,"D31","Tampon sonu"); out(ws,"E31",f"={P}B35",DT)
lbl(ws,"B32","Bedel (KDV hariç)"); out(ws,"C32",f"={P}B39",TL); ws["C32"].font=F(bold=True)
lbl(ws,"D32","KDV"); out(ws,"E32",f"={P}B40",TL)
lbl(ws,"B33","Genel toplam"); out(ws,"C33",f"={P}B41",TL); ws["C33"].font=F(bold=True,color=BLUE)
lbl(ws,"D33","Günlük birim"); out(ws,"E33",f"={P}B38",TL)
lbl(ws,"B34","Ödeme planı")
out(ws,"C34",f'="İmza %"&TEXT({P}C43*100,"0")&": "&TEXT({P}D43,"#,##0.00")&" TL"')
out(ws,"D34",f'="UAT %"&TEXT({P}C44*100,"0")&": "&TEXT({P}D44,"#,##0.00")&" TL"')
out(ws,"E34",f'="GA %"&TEXT({P}C45*100,"0")&": "&TEXT({P}D45,"#,##0.00")&" TL"')
band(36,"KABUL KRİTERLERİ VE VARSAYIMLAR")
acc=["Kabul: 12 altın senaryoda birebir sonuç (sapma < 0,01 TL); test kapsamı ≥ %85; API p95 < 300 ms; pilot mutabakatında açıklanamayan fark 0.",
 "Varsayımlar: SGK resmi API'si yoktur — bildirilen SPEK verisi müşteri tarafından CSV/ekran ile sağlanır; altyapı müşteri sunucusunda (Hetzner) barındırılır; mevzuat değişikliği kapsam değişikliği sayılır.",
 "Hariçler: SGK/e-Devlet ekran otomasyonu, çok müşterili SaaS, mobil uygulama. Fikri haklar: kaynak kod müşteriye özel kullanım lisansıyla teslim edilir."]
for j,t in enumerate(acc):
    r=37+j; ws.merge_cells(f"B{r}:E{r}"); c=ws[f"B{r}"]; c.value=t; c.font=F(color=MUTED); c.alignment=WRAP
    ws.row_dimensions[r].height=34
lbl(ws,"B41","Yüklenici",bold=True); lbl(ws,"D41","Müşteri (İnşaat firması)",bold=True)
for cc in ("B43","C43","D43","E43"):
    ws[cc].border=Border(bottom=Side(style="thin",color=NAVY))
lbl(ws,"B44","Ad Soyad / İmza / Tarih",color=MUTED); lbl(ws,"D44","Ad Soyad / İmza / Tarih",color=MUTED)
widths(ws,{"A":2,"B":30,"C":32,"D":26,"E":30,"F":2})
ws.print_area="B2:E45"
ws.page_setup.paperSize=9; ws.page_setup.orientation="portrait"
ws.page_setup.fitToWidth=1; ws.page_setup.fitToHeight=1
ws.sheet_properties.pageSetUpPr=PageSetupProperties(fitToPage=True)
ws.page_margins=PageMargins(left=0.6,right=0.6,top=0.6,bottom=0.6)
ws.print_options.horizontalCentered=True

# ============ 5 PARAMETRELER ============
ws=sheet(wb,"PARAMETRELER",AMBER)
title(ws,"Parametreler — Tüm Değişkenler Burada","Pareto: müşteri formu sade kalır; her değişken bu sayfadadır. Web raporu bu değerleri canlı okur.")
sect(ws,4,"1) Hız (velocity) modeli — tamamen vibecoding")
vel=[("Geliştirici sayısı (insan)",1,"0","LİNEER: 3 kişi → süre ÷ 3. Kod yazmaz; spec/test/onay yapar."),
     ("AI hız çarpanı (kod üretiminde ×)",4.0,"0.0","AI'ın insana göre üretim hızı."),
     ("AI'ya devredilebilir iş payı (0-1)",0.60,"0.00","Kalan pay insana bağlı: inceleme, doğrulama, onay.")]
for j,(t,v,fmt,n) in enumerate(vel):
    r=5+j; lbl(ws,f"A{r}",t); inp(ws,f"B{r}",v,fmt); lbl(ws,f"C{r}",n,color=MUTED); ws.merge_cells(f"C{r}:E{r}")
lbl(ws,"A8","Amdahl süre oranı (otomatik)"); out(ws,"B8","=ROUND((1-B7)+B7/B6,3)","0.000")
lbl(ws,"C8","Süre oranı = insana bağlı pay + AI payı ÷ AI hızı. Varsayılan 0,550.",color=MUTED); ws.merge_cells("C8:E8")
sect(ws,10,"ECA kuralları (Olay-Koşul-Aksiyon) — velocity ayarı")
hdr_row(ws,11,["Kural","Aktif?","Etki","Uygulanan"])
eca=[("Altın senaryolar faz başında hazır → hızlan %10","HAYIR",0.90),
     ("CI kırmızı oranı > %20 → yavaşla %15","HAYIR",1.15),
     ("Mevzuat değişikliği (kapsam etkisi) → yavaşla %20","HAYIR",1.20),
     ("AI token bütçesi kısıldı → yavaşla %25","HAYIR",1.25),
     ("Paralel AI ajan orkestrasyonu → hızlan %15","HAYIR",0.85)]
for j,(t,a,e) in enumerate(eca):
    r=12+j
    lbl(ws,f"A{r}",t); inp(ws,f"B{r}",a); inp(ws,f"C{r}",e,"0.00")
    out(ws,f"D{r}",f'=IF(B{r}="EVET",C{r},1)',"0.00")
    dvx=DataValidation(type="list",formula1='"EVET,HAYIR"'); ws.add_data_validation(dvx); dvx.add(f"B{r}")
lbl(ws,"A17","ECA çarpanı"); out(ws,"D17","=ROUND(PRODUCT(D12:D16),3)","0.000"); ws["D17"].font=F(bold=True)
lbl(ws,"A18","ZAMAN ÇARPANI (nihai)",bold=True)
out(ws,"B18","=ROUND(B8*D17/B5,3)","0.000"); ws["B18"].font=F(bold=True,size=14,color=BLUE); ws["B18"].fill=PatternFill("solid",start_color=BLUE_L)
lbl(ws,"C18","Formül: Amdahl × ECA ÷ geliştirici. Varsayılan 0,550; 3 geliştiricide 0,183.",color=MUTED); ws.merge_cells("C18:E18")
sect(ws,20,"2) Modül taban eforları (insan-eşdeğeri adam-gün)")
mod_ef=[("Çekirdek altyapı & CI (zorunlu)",20),("M1 · Ruhsat & alan bazlı hesap",25),("M2 · İhale & sözleşme hesabı + kısmi araştırma",20),
        ("M3 · Aylık bildirim matrisi, İcmal & kota",30),("M4 · Mevzuat parametre yönetimi",15),("M5 · Google Sheets köprüsü & veri teyidi",18),
        ("M6 · Taşeron sicili, raporlar & KVKK",25)]
for j,(ad,g) in enumerate(mod_ef):
    r=21+j; lbl(ws,f"A{r}",ad); inp(ws,f"B{r}",g,'#,##0" a-g"')
lbl(ws,"A29","Seçili toplam efor (adam-gün)"); out(ws,"B29",f'={G}D22','#,##0" a-g"')
sect(ws,30,"3) Takvim")
lbl(ws,"A31","Efektif süre (gün, otomatik)"); out(ws,"B31","=MAX(10,ROUND(B29*B18,0))",'0" gün"'); ws["B31"].font=F(bold=True)
lbl(ws,"C31","Süre hesabı: efor × zaman çarpanı. Varsayılan: 153 × 0,550 = 84 gün.",color=MUTED); ws.merge_cells("C31:E31")
lbl(ws,"A33","Proje başlangıcı"); inp(ws,"B33",_dt.date(2026,7,20),DT)
lbl(ws,"A34","GA (otomatik)"); out(ws,"B34",f"={W}F12",DT); ws["B34"].font=F(bold=True,color=BLUE)
lbl(ws,"A35","Tampon sonu (+%15, otomatik)"); out(ws,"B35","=B34+MAX(3,ROUND(B31*0.15,0))",DT)
sect(ws,37,"4) Fiyatlama")
lbl(ws,"A38","Günlük birim ücret (adam-gün, KDV hariç)"); inp(ws,"B38",4000,TL)
lbl(ws,"A39","Bedel (KDV hariç, otomatik)"); out(ws,"B39","=ROUND(B29*B38,2)",TL); ws["B39"].font=F(bold=True)
lbl(ws,"A40","KDV (%20)"); out(ws,"B40","=ROUND(B39*0.2,2)",TL)
lbl(ws,"A41","Genel toplam"); out(ws,"B41","=B39+B40",TL); ws["B41"].font=F(bold=True,color=BLUE)
lbl(ws,"A42","Ödeme planı",bold=True)
for j,(ad,pct) in enumerate([("İmza (avans)",0.30),("UAT kabulü",0.40),("GA teslimi",0.30)]):
    r=43+j; lbl(ws,f"A{r}",ad); inp(ws,f"C{r}",pct,PCT); out(ws,f"D{r}",f"=ROUND($B$39*C{r},2)",TL)
sect(ws,47,"5) Mevzuat mini parametreler (ürün bağlamı; web mevzuat bölümünü besler)")
hdr_row(ws,48,["Dönem başlangıcı","Fark prim oranı","Asgari ücret (aylık brüt)","Not"])
don=[(_dt.date(2024,1,1),0.345,20002.50,"KVSK %2"),(_dt.date(2024,9,1),0.3475,20002.50,"7524: KVSK %2,25"),
     (_dt.date(2025,1,1),0.3475,26005.50,"2025 ücreti"),(_dt.date(2026,1,1),0.3575,33030.00,"7566: MYÖ %21, tavan 9 kat")]
for j,(d,o,u,n) in enumerate(don):
    r=49+j
    out(ws,f"A{r}",d,DT); out(ws,f"B{r}",o,PCT); out(ws,f"C{r}",u,TL); lbl(ws,f"D{r}",n,color=MUTED)
hdr_row(ws,54,["Yapı sınıfı","2025 (TL/m²)","2026 (TL/m²)",""])
mal=[("III-A",17100,19800),("III-B",18200,21050),("III-C",19150,23400),("IV-A",21500,26450),
     ("IV-B",27500,33900),("IV-C",32600,40500),("V-A",34400,42350),("V-D",43400,53500)]
for j,(s,a,b) in enumerate(mal):
    r=55+j; lbl(ws,f"A{r}",s); out(ws,f"B{r}",a,'#,##0'); out(ws,f"C{r}",b,'#,##0')
lbl(ws,"A64","Kaynak: RG 31.01.2025/32799 · RG 03.02.2026/33157 · SGK Genelge 2026/16 · 7566 s.K. Eski tablolardaki 0,345/0,375/0,385 çarpanları günceldışıdır.",color=MUTED)
ws.merge_cells("A64:E65"); ws["A64"].alignment=WRAP
widths(ws,{"A":46,"B":24,"C":22,"D":18,"E":40})
ws.freeze_panes="A5"

# ============ 6 WEB_PARAM (TÜMÜ METİN PROTOKOLÜ) ============
ws=sheet(wb,"WEB_PARAM",GRAY_H)
ws["A1"]="TIP"; ws["B1"]="AD"; ws["C1"]="DEGER"; ws["D1"]="EK1"; ws["E1"]="EK2"; ws["F1"]="EK3"
for c in "ABCDEF": ws[f"{c}1"].font=F(bold=True)
ws["H1"]="Web beslemesi. TÜM değerler metindir (gviz tip karışması önlemi). Elle değiştirmeyin."
ws["H1"].font=F(color=MUTED)
S=lambda ref: f'=""&{ref}'                      # metin zorla
N=lambda ref,dec=3: f'=""&ROUND({ref},{dec})'   # sayı→metin
D=lambda ref: f'=TEXT({ref},"yyyy-mm-dd")'      # tarih→ISO metin
kv=[("surum",'="4.1"'),("guncelleme",'=TEXT(NOW(),"yyyy-mm-dd hh:mm")'),
    ("dev",N(P+"B5",0)),("ai_hiz",N(P+"B6",1)),("ai_pay",N(P+"B7",2)),("amdahl",N(P+"B8",3)),
    ("eca",N(P+"D17",3)),("carpan",N(P+"B18",3)),
    ("efor",N(P+"B29",0)),("sure",N(P+"B31",0)),
    ("baslangic",D(P+"B33")),("ga",D(P+"B34")),("tampon",D(P+"B35")),
    ("gunluk_ucret",N(P+"B38",2)),("bedel",N(P+"B39",2)),("kdv",N(P+"B40",2)),("toplam",N(P+"B41",2)),
    ("asgari_ucret_2026",N(P+"C52",2)),("fark_prim_2026",N(P+"B52",4))]
for j,(k,v) in enumerate(kv):
    r=2+j; ws[f"A{r}"]="KV"; ws[f"B{r}"]=k; ws[f"C{r}"]=v
for j in range(7):
    r=23+j; gr=14+j
    ws[f"A{r}"]="MODUL"; ws[f"B{r}"]=S(G+f"B{gr}"); ws[f"C{r}"]=S(G+f"C{gr}"); ws[f"D{r}"]=N(G+f"D{gr}",0); ws[f"E{r}"]=S(G+f"E{gr}")
for j in range(6):
    r=32+j; wr=5+j
    ws[f"A{r}"]="FAZ"; ws[f"B{r}"]=S(W+f"A{wr}"); ws[f"C{r}"]=N(W+f"C{wr}",2); ws[f"D{r}"]=N(W+f"D{wr}",0)
    ws[f"E{r}"]=D(W+f"E{wr}"); ws[f"F{r}"]=D(W+f"F{wr}")
for j in range(4):
    r=40+j; pr=49+j
    ws[f"A{r}"]="DONEM"; ws[f"B{r}"]=D(P+f"A{pr}"); ws[f"C{r}"]=N(P+f"B{pr}",4); ws[f"D{r}"]=N(P+f"C{pr}",2)
for j in range(8):
    r=46+j; pr=55+j
    ws[f"A{r}"]="MALIYET"; ws[f"B{r}"]=S(P+f"A{pr}"); ws[f"C{r}"]=N(P+f"B{pr}",0); ws[f"D{r}"]=N(P+f"C{pr}",0)
tk=[("form_no",S("'TEKLİF FORMU'!C6")),("tarih",D("'TEKLİF FORMU'!E6")),("gecerlilik",S("'TEKLİF FORMU'!E7")),
    ("musteri",S("'TEKLİF FORMU'!C7")),("yuklenici",S("'TEKLİF FORMU'!C8")),("proje",S("'TEKLİF FORMU'!C9")),
    ("odeme1",S("'TEKLİF FORMU'!C34")),("odeme2",S("'TEKLİF FORMU'!D34")),("odeme3",S("'TEKLİF FORMU'!E34"))]
for j,(k,v) in enumerate(tk):
    r=56+j; ws[f"A{r}"]="TEKLIF"; ws[f"B{r}"]=k; ws[f"C{r}"]=v
widths(ws,{"A":10,"B":34,"C":40,"D":30,"E":30,"F":16,"H":70})

wb.save("/sessions/stoic-bold-goldberg/mnt/outputs/hrmshakedis_projelendirme_v4.xlsx")
print("saved v4:",wb.sheetnames)

import openpyxl
wb2=openpyxl.load_workbook("/sessions/stoic-bold-goldberg/mnt/outputs/hrmshakedis_projelendirme_v4.xlsx")
bad=["frappe","rönesans","ronesans","erpnext","asgari işçilik hesap / teklif"]
hits=[]
for wsx in wb2.worksheets:
    for row in wsx.iter_rows():
        for c in row:
            if isinstance(c.value,str):
                low=c.value.lower()
                for b in bad:
                    if b in low: hits.append((wsx.title,c.coordinate,b))
print("yasakli:",hits[:10] if hits else "TEMIZ")
